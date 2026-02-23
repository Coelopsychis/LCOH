from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, Tuple, Any

import openpyxl
import numpy as np

from .inputs import ScalarInputs
from .timeseries import to_hourly_series, HourlySeries


YELLOW_RGB = "FFFFFF00"  # #FFFF00 in ARGB (openpyxl stores ARGB)


@dataclass(frozen=True)
class ExcelCase:
    scalars: ScalarInputs
    ts: HourlySeries
    # reference outputs from Excel (saved values)
    ref_sheet2: Dict[str, float]


def _cell_is_yellow(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    # Many files store colors as ARGB
    return (fg.type == "rgb") and (fg.rgb == YELLOW_RGB)


def load_case_from_excel(path: str) -> ExcelCase:
    """
    Load a minimal runnable case from your Excel workbook.

    What we read right now:
      - Scalar inputs from Sheet '2. Ein- und Ausgaben' column C (yellow cells)
      - Time series:
          Sheet '8. Marktpreise': column B (day-ahead) rows 7..8766
          Sheet '9. Energiedaten': column P (standort1 own) rows 9..8768
                                 column AC (standort2 own) rows 9..8768
      - Reference outputs: Sheet 2 F3, F4, F7..F16 (cached values)
      - kWh/kg factor: Sheet '3. Nebenrechnungen'!C6
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws2 = wb["2. Ein- und Ausgaben"]
    ws3 = wb["3. Nebenrechnungen"]
    ws8 = wb["8. Marktpreise"]
    ws9 = wb["9. Energiedaten"]

    # --- Collect yellow inputs in sheet2 column C (mostly for later mapping) ---
    yellow_inputs: Dict[str, Any] = {}
    for row in range(1, ws2.max_row + 1):
        cell = ws2[f"C{row}"]
        if _cell_is_yellow(cell):
            yellow_inputs[f"C{row}"] = cell.value

    # Minimal mapping to ScalarInputs based on the known layout of your workbook.
    # If any of these addresses shift, we will make a robust label-based mapper later.
    commissioning_year = int(ws2["C5"].value)
    ely_kW = float(ws2["C6"].value)
    min_load = float(ws2["C10"].value)

    # PPA sizes (kW)
    ppa1_kW = float(ws2["C67"].value) if ws2["C67"].value is not None else 0.0
    ppa2_kW = float(ws2["C68"].value) if ws2["C68"].value is not None else 0.0

    # Spot threshold (EUR/MWh)
    spot_threshold = float(ws2["C76"].value) if ws2["C76"].value is not None else 0.0

    # Avg efficiency from sheet3 C29 (Excel uses that in F12)
    avg_eff = float(ws3["C29"].value)

    scalars = ScalarInputs(
        commissioning_year=commissioning_year,
        ely_kW=ely_kW,
        system_min_load=min_load,
        system_kW_override=None,  # start simple; can be replaced by Excel-derived system_kW later
        spot_enabled=True,
        spot_price_threshold_eur_per_mwh=spot_threshold,
        ppa1_enabled=True,
        ppa2_enabled=True,
        ppa1_kW=ppa1_kW,
        ppa2_kW=ppa2_kW,
        baseload_enabled=bool(ws2["C63"].value not in (None, 0, 0.0)),
        baseload_kW=float(ws2["C63"].value or 0.0),
        avg_efficiency=avg_eff,
    )

    # --- Time series ---
    # Day-ahead prices: sheet8 col B from row 7 for 8760 hours
    day_ahead = [ws8[f"B{r}"].value for r in range(7, 7 + 8760)]
    day_ahead = [0.0 if v is None else float(v) for v in day_ahead]

    # Own energy inputs:
    # sheet9 col P and AC from row 9 for 8760 hours
    pv1 = [ws9[f"P{r}"].value for r in range(9, 9 + 8760)]
    wind2 = [ws9[f"AC{r}"].value for r in range(9, 9 + 8760)]
    pv1 = [0.0 if v is None else float(v) for v in pv1]
    wind2 = [0.0 if v is None else float(v) for v in wind2]

    ts = to_hourly_series(pv1, wind2, day_ahead)

    # --- Reference outputs from Sheet2 ---
    ref = {}
    for addr in ["F3", "F4"] + [f"F{i}" for i in range(7, 17)]:
        v = ws2[addr].value
        ref[addr] = float(v) if v is not None else float("nan")

    return ExcelCase(
        scalars=scalars,
        ts=ts,
        ref_sheet2=ref,
    )


def load_kwh_per_kg_factor(path: str) -> float:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws3 = wb["3. Nebenrechnungen"]
    v = ws3["C6"].value
    return float(v) if v is not None else 0.0